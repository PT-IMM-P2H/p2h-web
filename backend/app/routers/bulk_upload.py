"""
Bulk upload endpoints for users and vehicles
"""
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, status
from sqlalchemy.orm import Session
import pandas as pd
import io
from typing import List
from datetime import datetime

from app.database import get_db
from app.dependencies import require_admin
from app.models.user import User, UserRole, UserKategori
from app.models.vehicle import Vehicle
from app.schemas.bulk_upload import BulkUploadResponse, BulkUploadError
from app.utils.password import hash_password
from app.utils.response import base_response
from app.repositories.vehicle_type_repository import VehicleTypeRepository

router = APIRouter(
    prefix="/bulk-upload",
    tags=["Bulk Upload"],
    dependencies=[Depends(require_admin)]
)


@router.post("/users", response_model=dict)
async def bulk_upload_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Bulk upload users from Excel file
    
    Expected columns:
    - Email (optional)
    - Nama Lengkap * (required)
    - Nomor Telepon * (required)
    - Tanggal Lahir (YYYY-MM-DD) (optional)
    - Role * (required: superadmin/admin/user)
    - Kategori * (required: PT/Travel)
    - Department (optional)
    - Position (optional)
    - Work Status (optional)
    
    Returns:
    - success_count: Number of successfully imported users
    - error_count: Number of failed rows
    - errors: List of errors with row numbers and messages
    """
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File harus berformat Excel (.xlsx atau .xls)"
        )
    
    try:
        # Read Excel file with converters to force phone number as string and preserve leading zeros
        contents = await file.read()
        
        # Custom converter function to preserve leading zeros in phone numbers
        def phone_converter(value):
            if pd.isna(value):
                return ''
            # Convert to string and remove any decimal points from float conversion
            s = str(value).strip()
            if '.' in s and s.replace('.', '').isdigit():
                # Remove decimal point if it's a whole number (e.g., "81234567890.0" -> "81234567890")
                s = str(int(float(s)))
            # Add leading zero if missing for Indonesian numbers (starts with 8 and has 10+ digits)
            if s.isdigit() and len(s) >= 10 and not s.startswith('0') and not s.startswith('+'):
                s = '0' + s
            return s
        
        df = pd.read_excel(
            io.BytesIO(contents),
            converters={'Nomor Telepon *': phone_converter}  # Apply converter to preserve phone format
        )
        
        # Normalize column names (remove spaces, lowercase)
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        # Map Excel column names to expected format
        # After normalization: "Nama Lengkap *" -> "nama_lengkap_*"
        column_mapping = {
            'email_*': 'email',
            'nama_lengkap_*': 'full_name',
            'nomor_telepon_*': 'phone_number',
            'tanggal_lahir_(dd-mm-yyyy)': 'birth_date',
            'tanggal_lahir_(yyyy-mm-dd)': 'birth_date',  # Support both formats
            'role_*': 'role',
            'kategori_*': 'kategori',
            'department_*': 'department',
            'position_*': 'position',
            'status_kerja_*': 'work_status',
            'nama_perusahaan_*': 'company'
        }
        
        df.rename(columns=column_mapping, inplace=True)
        
        success_count = 0
        reactivated_count = 0
        errors: List[BulkUploadError] = []
        
        # Preload all lookup data to avoid per-row queries
        from app.models.user import Department, Position, WorkStatus
        
        # Get all existing emails and phones for duplicate check (ACTIVE only)
        existing_emails = set(
            email[0].lower() for email in db.query(User.email).filter(
                User.email != None,
                User.is_active == True
            ).all()
        )
        existing_phones = set(
            phone[0] for phone in db.query(User.phone_number).filter(
                User.is_active == True
            ).all()
        )
        
        # Get soft-deleted users by phone for potential reactivation
        soft_deleted_by_phone = {
            u.phone_number: u for u in db.query(User).filter(
                User.is_active == False
            ).all()
        }
        
        # Preload departments, positions, work_statuses, companies as lookup dicts
        dept_lookup = {
            d.nama_department.lower(): d.id 
            for d in db.query(Department).filter(Department.deleted_at == None).all()
        }
        pos_lookup = {
            p.nama_posisi.lower(): p.id 
            for p in db.query(Position).filter(Position.deleted_at == None).all()
        }
        ws_lookup = {
            w.nama_status.lower(): w.id 
            for w in db.query(WorkStatus).filter(WorkStatus.deleted_at == None).all()
        }
        
        # Add Company lookup
        from app.models.user import Company
        company_lookup = {
            c.nama_perusahaan.lower(): c.id 
            for c in db.query(Company).filter(Company.deleted_at == None).all()
        }
        
        # Track new emails/phones added in this batch to prevent duplicates within file
        batch_emails = set()
        batch_phones = set()
        
        # Collect users to insert
        users_to_insert = []
        
        # Process each row
        for idx, row in df.iterrows():
            row_num = idx + 2  # Excel row number (1-indexed + header)
            
            try:
                # Skip completely empty rows
                if pd.isna(row.get('phone_number')) and pd.isna(row.get('full_name')):
                    continue
                
                # Validate required fields (email sekarang optional)
                # Email validation dihapus karena sudah nullable=True di model
                
                if pd.isna(row.get('full_name')) or not row.get('full_name'):
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='full_name',
                        message='Nama lengkap wajib diisi',
                        data=row.to_dict()
                    ))
                    continue
                
                if pd.isna(row.get('phone_number')) or not row.get('phone_number'):
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='phone_number',
                        message='Nomor telepon wajib diisi',
                        data=row.to_dict()
                    ))
                    continue
                
                if pd.isna(row.get('role')) or not row.get('role'):
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='role',
                        message='Role wajib diisi',
                        data=row.to_dict()
                    ))
                    continue
                
                if pd.isna(row.get('kategori')) or not row.get('kategori'):
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='kategori',
                        message='Kategori wajib diisi',
                        data=row.to_dict()
                    ))
                    continue
                
                # Process email
                email = None
                if not pd.isna(row.get('email')) and row.get('email'):
                    email = str(row['email']).strip().lower()
                    if email in existing_emails or email in batch_emails:
                        errors.append(BulkUploadError(
                            row=row_num,
                            field='email',
                            message=f'Email {email} sudah terdaftar pada user aktif',
                            data=row.to_dict()
                        ))
                        continue
                
                # Process phone
                phone = str(row['phone_number']).strip()
                if phone.startswith("'"):
                    phone = phone[1:]
                
                # Check if phone already exists in active users
                if phone in existing_phones or phone in batch_phones:
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='phone_number',
                        message=f'Nomor telepon {phone} sudah terdaftar pada user aktif',
                        data=row.to_dict()
                    ))
                    continue
                
                # Validate role
                role_str = str(row['role']).strip().lower()
                try:
                    role = UserRole(role_str)
                except ValueError:
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='role',
                        message=f'Role tidak valid: {role_str}. Gunakan: superadmin, admin, atau user',
                        data=row.to_dict()
                    ))
                    continue
                
                # Validate kategori
                kategori_str = str(row['kategori']).strip().upper()
                if kategori_str == 'PT':
                    kategori_str = 'IMM'  # Map PT to IMM
                try:
                    kategori = UserKategori(kategori_str)
                except ValueError:
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='kategori',
                        message=f'Kategori tidak valid: {kategori_str}. Gunakan: PT atau Travel',
                        data=row.to_dict()
                    ))
                    continue
                
                # Parse birth_date if provided (supports multiple formats)
                birth_date = None
                if not pd.isna(row.get('birth_date')):
                    try:
                        if isinstance(row['birth_date'], str):
                            date_str = row['birth_date'].strip()
                            # Try different date formats
                            for fmt in ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%y', '%d/%m/%y']:
                                try:
                                    birth_date = datetime.strptime(date_str, fmt).date()
                                    break
                                except ValueError:
                                    continue
                            if birth_date is None:
                                raise ValueError(f"Cannot parse date: {date_str}")
                        else:
                            birth_date = pd.to_datetime(row['birth_date']).date()
                    except:
                        errors.append(BulkUploadError(
                            row=row_num,
                            field='birth_date',
                            message='Format tanggal lahir tidak valid. Gunakan DD-MM-YYYY',
                            data=row.to_dict()
                        ))
                        continue
                
                # Lookup IDs using preloaded dicts (much faster)
                department_id = None
                position_id = None
                work_status_id = None
                company_id = None
                
                if not pd.isna(row.get('department')) and row.get('department'):
                    dept_name = str(row['department']).strip().lower()
                    department_id = dept_lookup.get(dept_name)
                
                if not pd.isna(row.get('position')) and row.get('position'):
                    pos_name = str(row['position']).strip().lower()
                    position_id = pos_lookup.get(pos_name)
                
                if not pd.isna(row.get('work_status')) and row.get('work_status'):
                    ws_name = str(row['work_status']).strip().lower()
                    work_status_id = ws_lookup.get(ws_name)
                
                if not pd.isna(row.get('company')) and row.get('company'):
                    company_name = str(row['company']).strip().lower()
                    company_id = company_lookup.get(company_name)
                
                # Check if there's a soft-deleted user with same phone number
                existing_deleted_user = soft_deleted_by_phone.get(phone)
                
                if existing_deleted_user:
                    # Reactivate and update the existing soft-deleted user
                    existing_deleted_user.is_active = True
                    existing_deleted_user.full_name = str(row['full_name']).strip()
                    existing_deleted_user.email = email
                    existing_deleted_user.birth_date = birth_date
                    existing_deleted_user.role = role
                    existing_deleted_user.kategori_pengguna = kategori
                    existing_deleted_user.department_id = department_id
                    existing_deleted_user.position_id = position_id
                    existing_deleted_user.work_status_id = work_status_id
                    existing_deleted_user.company_id = company_id
                    existing_deleted_user.updated_at = datetime.utcnow()
                    # Don't reset password on reactivation
                    
                    reactivated_count += 1
                    success_count += 1
                    
                    # Remove from soft_deleted dict to prevent duplicate reactivation
                    del soft_deleted_by_phone[phone]
                else:
                    # Generate default password for new users
                    if birth_date:
                        default_password = birth_date.strftime('%d%m%Y')
                    else:
                        default_password = 'P@ssw0rd123'
                    
                    password_hash = hash_password(default_password)
                    
                    # Create new user object
                    user = User(
                        email=email,
                        full_name=str(row['full_name']).strip(),
                        phone_number=phone,
                        password_hash=password_hash,
                        birth_date=birth_date,
                        role=role,
                        kategori_pengguna=kategori,
                        department_id=department_id,
                        position_id=position_id,
                        work_status_id=work_status_id,
                        company_id=company_id,
                        is_active=True
                    )
                    
                    users_to_insert.append(user)
                    success_count += 1
                
                # Track this batch
                if email:
                    batch_emails.add(email)
                batch_phones.add(phone)
                
            except Exception as e:
                errors.append(BulkUploadError(
                    row=row_num,
                    field=None,
                    message=f'Error tidak terduga: {str(e)}',
                    data=row.to_dict() if hasattr(row, 'to_dict') else None
                ))
        
        # Bulk insert new users
        if users_to_insert:
            db.bulk_save_objects(users_to_insert)
        
        # Commit all changes (reactivations + new inserts)
        db.commit()
        
        # Prepare response
        response_data = BulkUploadResponse(
            success_count=success_count,
            error_count=len(errors),
            errors=errors,
            total_rows=len(df)
        )
        
        # Convert to dict and handle Timestamp serialization
        response_dict = response_data.model_dump()
        if 'errors' in response_dict:
            for err in response_dict['errors']:
                if 'data' in err and err['data']:
                    for key, value in err['data'].items():
                        # Convert pandas Timestamp to string
                        if pd.notna(value) and hasattr(value, 'isoformat'):
                            err['data'][key] = value.isoformat()
                        elif pd.isna(value):
                            err['data'][key] = None
        
        return base_response(
            message=f"Import selesai: {success_count} berhasil, {len(errors)} gagal",
            payload=response_dict,
            status_code=status.HTTP_200_OK
        )
        
    except pd.errors.EmptyDataError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File Excel kosong"
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error saat memproses file: {str(e)}"
        )


@router.get("/templates/users")
async def download_users_template(db: Session = Depends(get_db)):
    """Download Excel template for bulk user upload - auto-generated with latest master data"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from app.models.user import Company, Department, Position, WorkStatus
    
    # Fetch fresh master data from database
    companies = [c.nama_perusahaan for c in db.query(Company).filter(Company.deleted_at == None).all()]
    departments = [d.nama_department for d in db.query(Department).filter(Department.deleted_at == None).all()]
    positions = [p.nama_posisi for p in db.query(Position).filter(Position.deleted_at == None).all()]
    work_statuses = [w.nama_status for w in db.query(WorkStatus).filter(WorkStatus.deleted_at == None).all()]
    roles = [e.value for e in UserRole]
    categories = [e.value for e in UserKategori]
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Pengguna"
    
    headers = ["Nama Lengkap *", "Nomor Telepon *", "Email *", "Tanggal Lahir (DD-MM-YYYY)", 
               "Nama Perusahaan *", "Kategori *", "Department *", "Position *", "Status Kerja *", "Role *"]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Set headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 22
    
    # Reference sheet
    ws_ref = wb.create_sheet("Data Referensi")
    
    def populate_ref(col, title, values):
        ws_ref[f'{col}1'] = title
        ws_ref[f'{col}1'].font = Font(bold=True, color="4472C4", name="Arial")
        for i, val in enumerate(values, 2):
            ws_ref[f'{col}{i}'] = val
        ws_ref.column_dimensions[col].width = 25
        return len(values)
    
    len_roles = populate_ref('A', 'Role', roles)
    len_cat = populate_ref('B', 'Kategori', categories)
    len_dept = populate_ref('C', 'Department', departments)
    len_pos = populate_ref('D', 'Position', positions)
    len_status = populate_ref('E', 'Status Kerja', work_statuses)
    len_comp = populate_ref('F', 'Nama Perusahaan', companies)
    
    # Add data validations
    if len_roles > 0:
        dv = DataValidation(type="list", formula1=f"='Data Referensi'!$A$2:$A${len_roles+1}", allow_blank=False)
        dv.error = 'Pilih dari daftar Role yang tersedia'
        dv.errorTitle = 'Invalid Role'
        ws.add_data_validation(dv)
        dv.add('J2:J1000')
    
    if len_cat > 0:
        dv = DataValidation(type="list", formula1=f"='Data Referensi'!$B$2:$B${len_cat+1}", allow_blank=False)
        dv.error = 'Pilih dari daftar Kategori yang tersedia'
        dv.errorTitle = 'Invalid Kategori'
        ws.add_data_validation(dv)
        dv.add('F2:F1000')
    
    if len_dept > 0:
        dv = DataValidation(type="list", formula1=f"='Data Referensi'!$C$2:$C${len_dept+1}", allow_blank=False)
        dv.error = 'Pilih dari daftar Department yang tersedia'
        dv.errorTitle = 'Invalid Department'
        ws.add_data_validation(dv)
        dv.add('G2:G1000')
    
    if len_pos > 0:
        dv = DataValidation(type="list", formula1=f"='Data Referensi'!$D$2:$D${len_pos+1}", allow_blank=False)
        dv.error = 'Pilih dari daftar Position yang tersedia'
        dv.errorTitle = 'Invalid Position'
        ws.add_data_validation(dv)
        dv.add('H2:H1000')
    
    if len_status > 0:
        dv = DataValidation(type="list", formula1=f"='Data Referensi'!$E$2:$E${len_status+1}", allow_blank=False)
        dv.error = 'Pilih dari daftar Status Kerja yang tersedia'
        dv.errorTitle = 'Invalid Status Kerja'
        ws.add_data_validation(dv)
        dv.add('I2:I1000')
    
    if len_comp > 0:
        dv = DataValidation(type="list", formula1=f"='Data Referensi'!$F$2:$F${len_comp+1}", allow_blank=False)
        dv.error = 'Pilih dari daftar Perusahaan yang tersedia'
        dv.errorTitle = 'Invalid Perusahaan'
        ws.add_data_validation(dv)
        dv.add('E2:E1000')
    
    # Add empty row with highlighting for mandatory fields
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = ""
        cell.border = border
        if "*" in headers[col_num - 1]:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Format phone number column as TEXT
    for row in range(2, 1001):
        ws.cell(row=row, column=2).number_format = '@'
    
    ws.freeze_panes = 'A2'
    
    # Save to stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=template_data_pengguna.xlsx'}
    )


@router.get("/templates/vehicles")
async def download_vehicles_template(db: Session = Depends(get_db)):
    """Download Excel template for bulk vehicle upload - auto-generated with latest master data"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from app.models.vehicle import VehicleType, ShiftType, UnitKategori
    
    # Fetch fresh master data
    vehicle_types = [e.value for e in VehicleType]
    categories = [e.value for e in UnitKategori]
    shift_types = [e.value for e in ShiftType]
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Kendaraan"
    
    headers = ["Nomor Polisi *", "Nomor Lambung", "Lokasi Kendaraan", "Tipe Kendaraan *", 
               "Kategori *", "Merek", "Tahun Pembuatan", "Tanggal Expired STNK (YYYY-MM-DD)", 
               "Tanggal Expired Pajak (YYYY-MM-DD)", "Tanggal Expired KIR (YYYY-MM-DD)", "Shift Type *"]
    
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Set headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 25
    
    # Reference sheet
    ws_ref = wb.create_sheet("Data Referensi")
    
    def populate_ref(col, title, values):
        ws_ref[f'{col}1'] = title
        ws_ref[f'{col}1'].font = Font(bold=True, color="70AD47", name="Arial")
        for i, val in enumerate(values, 2):
            ws_ref[f'{col}{i}'] = val
        ws_ref.column_dimensions[col].width = 25
        return len(values)
    
    len_types = populate_ref('A', 'Tipe Kendaraan', vehicle_types)
    len_cat = populate_ref('B', 'Kategori', categories)
    len_shift = populate_ref('C', 'Shift Type', shift_types)
    len_lokasi = populate_ref('D', 'Lokasi Kendaraan', ['Port', 'KM. 30'])
    
    # Add data validations
    if len_types > 0:
        dv = DataValidation(type="list", formula1=f"='Data Referensi'!$A$2:$A${len_types+1}", allow_blank=False)
        dv.error = 'Pilih dari daftar Tipe Kendaraan yang tersedia'
        dv.errorTitle = 'Invalid Tipe'
        ws.add_data_validation(dv)
        dv.add('D2:D1000')
    
    if len_cat > 0:
        dv = DataValidation(type="list", formula1=f"='Data Referensi'!$B$2:$B${len_cat+1}", allow_blank=False)
        dv.error = 'Pilih dari daftar Kategori yang tersedia'
        dv.errorTitle = 'Invalid Kategori'
        ws.add_data_validation(dv)
        dv.add('E2:E1000')
    
    if len_shift > 0:
        dv = DataValidation(type="list", formula1=f"='Data Referensi'!$C$2:$C${len_shift+1}", allow_blank=False)
        dv.error = 'Pilih dari daftar Shift yang tersedia'
        dv.errorTitle = 'Invalid Shift'
        ws.add_data_validation(dv)
        dv.add('K2:K1000')
    
    if len_lokasi > 0:
        dv = DataValidation(type="list", formula1=f"='Data Referensi'!$D$2:$D${len_lokasi+1}", allow_blank=True)
        dv.error = 'Pilih Port atau KM. 30'
        dv.errorTitle = 'Invalid Lokasi'
        ws.add_data_validation(dv)
        dv.add('C2:C1000')
    
    # Add empty row with highlighting for mandatory fields
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = ""
        cell.border = border
        if "*" in headers[col_num - 1]:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    ws.freeze_panes = 'A2'
    
    # Save to stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=template_data_kendaraan.xlsx'}
    )


@router.post("/vehicles", response_model=dict)
async def bulk_upload_vehicles(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Bulk upload vehicles from Excel file
    
    Expected columns:
    - Nomor Polisi * (required)
    - Nomor Lambung (optional)
    - Tipe Kendaraan * (required)
    - Kategori * (required: PT/Travel)
    - Tahun Pembuatan (optional)
    - Tanggal Expired STNK (YYYY-MM-DD) (optional)
    - Tanggal Expired KIR (YYYY-MM-DD) (optional)
    - Shift Type * (required: Shift/Non Shift/Long Shift)
    
    Returns:
    - success_count: Number of successfully imported vehicles
    - error_count: Number of failed rows
    - errors: List of errors with row numbers and messages
    """
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File harus berformat Excel (.xlsx atau .xls)"
        )
    
    try:
        # Read Excel file
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        column_mapping = {
            'nomor_polisi_*': 'nomor_polisi',
            'nomor_lambung': 'nomor_lambung',
            'tipe_kendaraan_*': 'tipe_kendaraan',
            'kategori_*': 'kategori',
            'tahun_pembuatan': 'tahun_pembuatan',
            'tanggal_expired_stnk_(yyyy-mm-dd)': 'expired_stnk',
            'tanggal_expired_kir_(yyyy-mm-dd)': 'expired_kir',
            'shift_type_*': 'shift_type'
        }
        
        df.rename(columns=column_mapping, inplace=True)
        
        success_count = 0
        errors: List[BulkUploadError] = []
        
        # Get all vehicle types from database for validation
        vehicle_type_repo = VehicleTypeRepository(db)
        active_types = vehicle_type_repo.get_active()
        valid_type_names = {vt.name: vt.name for vt in active_types}
        
        # Process each row
        for idx, row in df.iterrows():
            row_num = idx + 2
            
            try:
                # Skip completely empty rows
                if pd.isna(row.get('nomor_polisi')) and pd.isna(row.get('tipe_kendaraan')):
                    continue
                
                # Validate required fields
                if pd.isna(row.get('nomor_polisi')) or not row.get('nomor_polisi'):
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='nomor_polisi',
                        message='Nomor polisi wajib diisi',
                        data=row.to_dict()
                    ))
                    continue
                
                if pd.isna(row.get('tipe_kendaraan')) or not row.get('tipe_kendaraan'):
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='tipe_kendaraan',
                        message='Tipe kendaraan wajib diisi',
                        data=row.to_dict()
                    ))
                    continue
                
                if pd.isna(row.get('kategori')) or not row.get('kategori'):
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='kategori',
                        message='Kategori wajib diisi',
                        data=row.to_dict()
                    ))
                    continue
                
                if pd.isna(row.get('shift_type')) or not row.get('shift_type'):
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='shift_type',
                        message='Shift type wajib diisi',
                        data=row.to_dict()
                    ))
                    continue
                
                # Check for duplicate plat_nomor
                plat = str(row['nomor_polisi']).strip().upper()
                existing_vehicle = db.query(Vehicle).filter(Vehicle.plat_nomor == plat).first()
                if existing_vehicle:
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='nomor_polisi',
                        message=f'Nomor polisi {plat} sudah terdaftar',
                        data=row.to_dict()
                    ))
                    continue
                
                # Validate vehicle type against database
                type_str = str(row['tipe_kendaraan']).strip()
                if type_str not in valid_type_names:
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='tipe_kendaraan',
                        message=f'Tipe kendaraan tidak valid: {row["tipe_kendaraan"]}. Pilih dari: {", ".join(vt.name for vt in active_types)}',
                        data=row.to_dict()
                    ))
                    continue
                
                # Use the vehicle type name directly (it's stored as string/enum in current schema)
                vehicle_type_value = type_str
                
                # Validate kategori
                kategori_str = str(row['kategori']).strip().upper()
                if kategori_str == 'PT':
                    kategori_str = 'IMM'
                if kategori_str not in ['IMM', 'TRAVEL']:
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='kategori',
                        message=f'Kategori tidak valid: {row["kategori"]}. Gunakan: PT atau Travel',
                        data=row.to_dict()
                    ))
                    continue
                
                # Validate shift type
                shift_str = str(row['shift_type']).strip().lower()
                if 'shift 1' in shift_str or 'shift 2' in shift_str or 'shift 3' in shift_str or shift_str in ['1', '2', '3', 'shift']:
                    shift_type = 'shift'
                elif 'long' in shift_str:
                    shift_type = 'long_shift'
                elif 'non' in shift_str:
                    shift_type = 'non_shift'
                else:
                    errors.append(BulkUploadError(
                        row=row_num,
                        field='shift_type',
                        message=f'Shift type tidak valid: {row["shift_type"]}. Gunakan: Shift/Non Shift/Long Shift',
                        data=row.to_dict()
                    ))
                    continue
                
                # Parse dates if provided
                stnk_expiry = None
                kir_expiry = None
                
                if not pd.isna(row.get('expired_stnk')):
                    try:
                        if isinstance(row['expired_stnk'], str):
                            stnk_expiry = datetime.strptime(row['expired_stnk'], '%Y-%m-%d').date()
                        else:
                            stnk_expiry = pd.to_datetime(row['expired_stnk']).date()
                    except:
                        errors.append(BulkUploadError(
                            row=row_num,
                            field='expired_stnk',
                            message='Format tanggal STNK tidak valid. Gunakan YYYY-MM-DD',
                            data=row.to_dict()
                        ))
                        continue
                
                if not pd.isna(row.get('expired_kir')):
                    try:
                        if isinstance(row['expired_kir'], str):
                            kir_expiry = datetime.strptime(row['expired_kir'], '%Y-%m-%d').date()
                        else:
                            kir_expiry = pd.to_datetime(row['expired_kir']).date()
                    except:
                        errors.append(BulkUploadError(
                            row=row_num,
                            field='expired_kir',
                            message='Format tanggal KIR tidak valid. Gunakan YYYY-MM-DD',
                            data=row.to_dict()
                        ))
                        continue
                
                # Create vehicle
                vehicle = Vehicle(
                    plat_nomor=plat,
                    no_lambung=str(row['nomor_lambung']).strip() if not pd.isna(row.get('nomor_lambung')) else None,
                    vehicle_type=vehicle_type_value,
                    kategori_unit=kategori_str,
                    shift_type=shift_type,
                    stnk_expiry=stnk_expiry,
                    kir_expiry=kir_expiry,
                    is_active=True
                )
                
                db.add(vehicle)
                success_count += 1
                
            except Exception as e:
                errors.append(BulkUploadError(
                    row=row_num,
                    field=None,
                    message=f'Error tidak terduga: {str(e)}',
                    data=row.to_dict() if hasattr(row, 'to_dict') else None
                ))
        
        # Commit all successful inserts
        if success_count > 0:
            db.commit()
        
        # Prepare response
        response_data = BulkUploadResponse(
            success_count=success_count,
            error_count=len(errors),
            errors=errors,
            total_rows=len(df)
        )
        
        # Convert to dict and handle Timestamp serialization
        response_dict = response_data.model_dump()
        if 'errors' in response_dict:
            for err in response_dict['errors']:
                if 'data' in err and err['data']:
                    for key, value in err['data'].items():
                        # Convert pandas Timestamp to string
                        if pd.notna(value) and hasattr(value, 'isoformat'):
                            err['data'][key] = value.isoformat()
                        elif pd.isna(value):
                            err['data'][key] = None
        
        return base_response(
            message=f"Import selesai: {success_count} berhasil, {len(errors)} gagal",
            payload=response_dict,
            status_code=status.HTTP_200_OK
        )
        
    except pd.errors.EmptyDataError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File Excel kosong"
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error saat memproses file: {str(e)}"
        )
