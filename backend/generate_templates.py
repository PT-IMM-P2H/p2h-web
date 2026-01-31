"""
Generate Excel templates for bulk upload with data validation from Database
Run this script to create template files for users and vehicles
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os
import sys

# Ensure backend directory is in path to import app modules
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from app.database import SessionLocal
from app.models.user import Company, Department, Position, WorkStatus, UserRole, UserKategori
from app.models.vehicle import VehicleType, ShiftType, UnitKategori

# Create templates directory
templates_dir = os.path.join(os.path.dirname(__file__), 'templates')
os.makedirs(templates_dir, exist_ok=True)


def get_master_data():
    """Fetch all master data from database"""
    db = SessionLocal()
    try:
        data = {
            "companies": [c.nama_perusahaan for c in db.query(Company).filter(Company.deleted_at == None).all()],
            "departments": [d.nama_department for d in db.query(Department).filter(Department.deleted_at == None).all()],
            "positions": [p.nama_posisi for p in db.query(Position).filter(Position.deleted_at == None).all()],
            "work_statuses": [w.nama_status for w in db.query(WorkStatus).filter(WorkStatus.deleted_at == None).all()],
            "roles": [e.value for e in UserRole],
            "user_categories": [e.value for e in UserKategori],
            "vehicle_types": [e.value for e in VehicleType],
            "vehicle_categories": [e.value for e in UnitKategori],
            "shift_types": [e.value for e in ShiftType],
        }
        return data
    finally:
        db.close()


def create_users_template(master_data):
    """Generate template for bulk user upload with reference data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Pengguna"
    
    headers = [
        "Nama Lengkap *",
        "Nomor Telepon *",
        "Email *",
        "Tanggal Lahir (DD-MM-YYYY)",
        "Nama Perusahaan *",
        "Kategori *",
        "Department *",
        "Position *",
        "Status Kerja *",
        "Role *",
    ]
    
    # Style configuration
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 22
    
    # Create reference data sheet
    ws_ref = wb.create_sheet("Data Referensi")
    
    # Helper to populate reference column
    def populate_ref_col(col_letter, title, values):
        ws_ref[f'{col_letter}1'] = title
        ws_ref[f'{col_letter}1'].font = Font(bold=True, color="4472C4", name="Arial")
        for i, val in enumerate(values, 2):
            ws_ref[f'{col_letter}{i}'] = val
        return len(values)

    # Populate references
    len_roles = populate_ref_col('A', 'Role', master_data['roles'])
    len_kategori = populate_ref_col('B', 'Kategori', master_data['user_categories'])
    len_dept = populate_ref_col('C', 'Department', master_data['departments'])
    len_pos = populate_ref_col('D', 'Position', master_data['positions'])
    len_status = populate_ref_col('E', 'Status Kerja', master_data['work_statuses'])
    len_comp = populate_ref_col('F', 'Nama Perusahaan', master_data['companies'])
    
    # Adjust column widths in reference sheet
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_ref.column_dimensions[col].width = 25
    
    # Data validations using reference sheet
    # Role validation (column J)
    if len_roles > 0:
        role_validation = DataValidation(
            type="list",
            formula1=f"='Data Referensi'!$A$2:$A${len_roles+1}",
            allow_blank=False
        )
        role_validation.error = 'Pilih dari daftar Role yang tersedia'
        role_validation.errorTitle = 'Invalid Role'
        ws.add_data_validation(role_validation)
        role_validation.add('J2:J10000')
    
    # Kategori validation (column F)
    if len_kategori > 0:
        kategori_validation = DataValidation(
            type="list",
            formula1=f"='Data Referensi'!$B$2:$B${len_kategori+1}",
            allow_blank=False
        )
        kategori_validation.error = 'Pilih dari daftar Kategori yang tersedia'
        kategori_validation.errorTitle = 'Invalid Kategori'
        ws.add_data_validation(kategori_validation)
        kategori_validation.add('F2:F10000')
    
    # Department validation (column G)
    if len_dept > 0:
        dept_validation = DataValidation(
            type="list",
            formula1=f"='Data Referensi'!$C$2:$C${len_dept+1}",
            allow_blank=False
        )
        dept_validation.error = 'Pilih dari daftar Department yang tersedia'
        dept_validation.errorTitle = 'Invalid Department'
        ws.add_data_validation(dept_validation)
        dept_validation.add('G2:G10000')
    
    # Position validation (column H)
    if len_pos > 0:
        pos_validation = DataValidation(
            type="list",
            formula1=f"='Data Referensi'!$D$2:$D${len_pos+1}",
            allow_blank=False
        )
        pos_validation.error = 'Pilih dari daftar Position yang tersedia'
        pos_validation.errorTitle = 'Invalid Position'
        ws.add_data_validation(pos_validation)
        pos_validation.add('H2:H10000')
    
    # Status Kerja validation (column I)
    if len_status > 0:
        status_validation = DataValidation(
            type="list",
            formula1=f"='Data Referensi'!$E$2:$E${len_status+1}",
            allow_blank=False
        )
        status_validation.error = 'Pilih dari daftar Status Kerja yang tersedia'
        status_validation.errorTitle = 'Invalid Status Kerja'
        ws.add_data_validation(status_validation)
        status_validation.add('I2:I10000')
    
    # Nama Perusahaan validation (column E)
    if len_comp > 0:
        company_validation = DataValidation(
            type="list",
            formula1=f"='Data Referensi'!$F$2:$F${len_comp+1}",
            allow_blank=False
        )
        company_validation.error = 'Pilih dari daftar Perusahaan yang tersedia'
        company_validation.errorTitle = 'Invalid Perusahaan'
        ws.add_data_validation(company_validation)
        company_validation.add('E2:E10000')
    
    # Add empty row with highlighting for mandatory fields
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = ""
        cell.border = border
        if "*" in headers[col_num - 1]:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Format phone number column as TEXT
    for row in range(2, 1001):
        cell = ws.cell(row=row, column=2)
        cell.number_format = '@'
    
    ws.freeze_panes = 'A2'
    
    # Instructions sheet
    ws_inst = wb.create_sheet("Instruksi")
    instructions = [
        ["INSTRUKSI UPLOAD DATA PENGGUNA", ""],
        ["", ""],
        ["1. Format File:", ""],
        ["", "- Simpan file sebagai .xlsx (Excel)"],
        ["", "- Jangan ubah nama kolom header"],
        ["", ""],
        ["2. Kolom Wajib (*):", ""],
        ["", "- Nama Lengkap: Nama lengkap pengguna"],
        ["", "- Nomor Telepon: Ketik langsung dengan 0 di depan (contoh: 081234567890)"],
        ["", "  PENTING: Kolom sudah diformat TEXT, angka 0 tidak akan hilang"],
        ["", "- Email: Format email valid (contoh: user@example.com)"],
        ["", "- Nama Perusahaan: Pilih dari dropdown"],
        ["", "- Kategori: Pilih dari dropdown (IMM/TRAVEL)"],
        ["", "- Department: Pilih dari dropdown"],
        ["", "- Position: Pilih dari dropdown"],
        ["", "- Status Kerja: Pilih dari dropdown"],
        ["", "- Role: Pilih dari dropdown (superadmin/admin/user)"],
        ["", ""],
        ["3. Kolom Opsional:", ""],
        ["", "- Tanggal Lahir: Format DD-MM-YYYY (contoh: 15-01-1990)"],
        ["", ""],
        ["4. Data Referensi:", ""],
        ["", "- Lihat sheet 'Data Referensi' untuk daftar lengkap pilihan"],
        ["", "- Dropdown otomatis terhubung ke data referensi yang ada di database"],
        ["", ""],
        ["5. Catatan:", ""],
        ["", "- Email duplikat akan dilewati"],
        ["", "- Data invalid akan dilaporkan sebagai error"],
        ["", "- Password default: 'P@ssw0rd123' (wajib ganti saat login pertama)"],
    ]
    
    for row_num, row_data in enumerate(instructions, 1):
        ws_inst.cell(row=row_num, column=1).value = row_data[0]
        ws_inst.cell(row=row_num, column=2).value = row_data[1]
        if row_num == 1:
            ws_inst.cell(row=row_num, column=1).font = Font(size=14, bold=True, color="4472C4", name="Arial")
        else:
            ws_inst.cell(row=row_num, column=1).font = Font(name="Arial")
            ws_inst.cell(row=row_num, column=2).font = Font(name="Arial")
    
    ws_inst.column_dimensions['A'].width = 25
    ws_inst.column_dimensions['B'].width = 65
    
    filepath = os.path.join(templates_dir, 'template_data_pengguna.xlsx')
    wb.save(filepath)
    print(f"✅ Template pengguna created: {filepath}")


def create_vehicles_template(master_data):
    """Generate template for bulk vehicle upload with reference data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Kendaraan"
    
    headers = [
        "Nomor Polisi *",
        "Nomor Lambung",
        "Tipe Kendaraan *",
        "Kategori *",
        "Tahun Pembuatan",
        "Tanggal Expired STNK (YYYY-MM-DD)",
        "Tanggal Expired KIR (YYYY-MM-DD)",
        "Shift Type *"
    ]
    
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 25
    
    # Create reference data sheet
    ws_ref = wb.create_sheet("Data Referensi")
    
    def populate_ref_col(col_letter, title, values):
        ws_ref[f'{col_letter}1'] = title
        ws_ref[f'{col_letter}1'].font = Font(bold=True, color="70AD47", name="Arial")
        for i, val in enumerate(values, 2):
            ws_ref[f'{col_letter}{i}'] = val
        return len(values)

    len_types = populate_ref_col('A', 'Tipe Kendaraan', master_data['vehicle_types'])
    len_kat = populate_ref_col('B', 'Kategori', master_data['vehicle_categories'])
    len_shift = populate_ref_col('C', 'Shift Type', master_data['shift_types'])
    
    for col in ['A', 'B', 'C']:
        ws_ref.column_dimensions[col].width = 25
    
    # Data validations
    # Tipe Kendaraan validation (column C)
    if len_types > 0:
        type_validation = DataValidation(
            type="list",
            formula1=f"='Data Referensi'!$A$2:$A${len_types+1}",
            allow_blank=False
        )
        type_validation.error = 'Pilih dari daftar Tipe Kendaraan yang tersedia'
        type_validation.errorTitle = 'Invalid Tipe'
        ws.add_data_validation(type_validation)
        type_validation.add('C2:C10000')
    
    # Kategori validation (column D)
    if len_kat > 0:
        kategori_validation = DataValidation(
            type="list",
            formula1=f"='Data Referensi'!$B$2:$B${len_kat+1}",
            allow_blank=False
        )
        kategori_validation.error = 'Pilih dari daftar Kategori yang tersedia'
        kategori_validation.errorTitle = 'Invalid Kategori'
        ws.add_data_validation(kategori_validation)
        kategori_validation.add('D2:D10000')
    
    # Shift Type validation (column H)
    if len_shift > 0:
        shift_validation = DataValidation(
            type="list",
            formula1=f"='Data Referensi'!$C$2:$C${len_shift+1}",
            allow_blank=False
        )
        shift_validation.error = 'Pilih dari daftar Shift yang tersedia'
        shift_validation.errorTitle = 'Invalid Shift'
        ws.add_data_validation(shift_validation)
        shift_validation.add('H2:H10000')
    
    # Add empty row with highlighting
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = ""
        cell.border = border
        if "*" in headers[col_num - 1]:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    ws.freeze_panes = 'A2'
    
    # Instructions sheet
    ws_inst = wb.create_sheet("Instruksi")
    instructions = [
        ["INSTRUKSI UPLOAD DATA KENDARAAN", ""],
        ["", ""],
        ["1. Format File:", ""],
        ["", "- Simpan file sebagai .xlsx (Excel)"],
        ["", "- Jangan ubah nama kolom header"],
        ["", ""],
        ["2. Kolom Wajib (*):", ""],
        ["", "- Nomor Polisi: Format standar (contoh: KT 1234 AB)"],
        ["", "- Tipe Kendaraan: Pilih dari dropdown"],
        ["", "- Kategori: Pilih dari dropdown (IMM/TRAVEL)"],
        ["", "- Shift Type: Pilih dari dropdown"],
        ["", ""],
        ["3. Kolom Opsional:", ""],
        ["", "- Nomor Lambung: Nomor identifikasi internal"],
        ["", "- Tahun Pembuatan: 4 digit (contoh: 2020)"],
        ["", "- Tanggal Expired STNK/KIR: Format YYYY-MM-DD (contoh: 2025-12-31)"],
        ["", ""],
        ["4. Data Referensi:", ""],
        ["", "- Lihat sheet 'Data Referensi' untuk daftar lengkap pilihan"],
        ["", "- Dropdown otomatis terhubung ke data referensi di database"],
        ["", ""],
        ["5. Catatan:", ""],
        ["", "- Nomor polisi duplikat akan dilewati"],
        ["", "- Data invalid akan dilaporkan sebagai error"],
        ["", "- Tanggal expired untuk reminder otomatis"],
    ]
    
    for row_num, row_data in enumerate(instructions, 1):
        ws_inst.cell(row=row_num, column=1).value = row_data[0]
        ws_inst.cell(row=row_num, column=2).value = row_data[1]
        if row_num == 1:
            ws_inst.cell(row=row_num, column=1).font = Font(size=14, bold=True, color="70AD47", name="Arial")
        else:
            ws_inst.cell(row=row_num, column=1).font = Font(name="Arial")
            ws_inst.cell(row=row_num, column=2).font = Font(name="Arial")
    
    ws_inst.column_dimensions['A'].width = 30
    ws_inst.column_dimensions['B'].width = 65
    
    filepath = os.path.join(templates_dir, 'template_data_kendaraan.xlsx')
    wb.save(filepath)
    print(f"✅ Template kendaraan created: {filepath}")


if __name__ == "__main__":
    print("🔧 Connect to Database...")
    try:
        master_data = get_master_data()
        print("✅ Master data fetched successfully")
        
        print("🔧 Generating Excel templates...")
        create_users_template(master_data)
        create_vehicles_template(master_data)
        
        print("")
        print("✅ All templates generated successfully!")
        print(f"📁 Location: {templates_dir}")
    except Exception as e:
        print(f"❌ Error: {str(e)}")